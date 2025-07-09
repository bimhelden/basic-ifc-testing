import requests
import xml.etree.ElementTree as ET
import sys
import re
import csv
import json
import os
from github import Github
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

replace_id = False
download_csv = False
output_file = ""
error_codes = set()
missing_messages = set()

def fetch_file_from_github(local_file_path = "ChkXML-IFC4Add2TC1-DE.csv"):
    # URL of the raw file to be downloaded
    url = "https://raw.githubusercontent.com/bimhelden/basic-ifc-testing/main/Fehlerdokumentation/" + local_file_path
    print(f"Trying to download {url} ..")
    response = requests.get(url)

    # Check if the request was successful
    if response.status_code == 200:
        # Write the content to a local file
        with open(local_file_path, "wb") as file:
            file.write(response.content)
        print(f"  File downloaded successfully and saved to {local_file_path}")
    else:
        print(f"  Failed to download file. HTTP status code: {response.status_code}")

def read_csv_to_hash(csv_filename):
    data_hash = {}
    try:
        with open(csv_filename, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                key = row['WR']
                del row['WR']  # Remove the key from the row data
                data_hash[key] = row
        return data_hash
    except FileNotFoundError:
        print(f"Error: The file '{csv_filename}' was not found.")
        sys.exit(-1)
    except Exception as e:
        print(f"An unexpected error occurred when trying to read the translation file: {e}")
        sys.exit(-1)

# Create or update a JSON file with statistical data and information about missing error codes
def update_statistics(json_file_path, keys_to_update, values_to_update):
    # Download statistics from Github
    try:
        fetch_file_from_github(json_file_path)
    except Exception as e:
        print(f"Download of the statistics file failed with the following error code: {e}")

    # Check if the file exists
    if os.path.exists(json_file_path):
        # Read the existing file
        with open(json_file_path, 'r') as file:
            data = json.load(file)
    else:
        # Initialize with the default structure if the file does not exist
        data = {
            "Number of checks": 0,
            "Statistics": {},
            "Unknown": []
        }

    # Increment the "Number of checks"
    if "Number of checks" in data:
        data["Number of checks"] += 1
    else:
        data["Number of checks"] = 1

    # Update the "Statistics" with the keys
    for key in keys_to_update:
        if key in data["Statistics"]:
            data["Statistics"][key] += 1
        else:
            data["Statistics"][key] = 1

    # Sort the Statistics dictionary by key
    data["Statistics"] = dict(sorted(data["Statistics"].items()))

    # Update the "Unknown" set with the values
    if "Unknown" in data:
        unknown_set = set(data["Unknown"])
    else:
        unknown_set = set()

    for value in values_to_update:
        unknown_set.add(value)

    data["Unknown"] = sorted(list(unknown_set))

    # Write the updated data back to the file
    with open(json_file_path, 'w') as file:
        json.dump(data, file, indent=4)

def replace_substring(text):
    # Define the regular expression pattern to match substrings starting with #
    pattern = r'#\d+'

    # Define the replacement string (by unknown id)
    replacement = '#uid'

    # Use re.sub() to replace all occurrences of the pattern with the replacement string
    return re.sub(pattern, replacement, text)

def sort_xml(input_file, output_file, replace_id, translation_hash = None, download_csv = False):
    # Parse the XML file
    try:
        tree = ET.parse(input_file)
    except FileNotFoundError:
        print(f"Error: Input file '{input_file}' was not found.")
        sys.exit(-1)

    root = tree.getroot()

    # Replace substrings in XML content
    if replace_id is True:
        for elem in root.iter():
            # Check if the element's text contains the substring to be replaced
            if elem.text is not None:
                elem.text = replace_substring(elem.text)

    # Sort MVDObjectStatistic elements based on ObjectName
    object_statistics_set = root.find('.//MVDObjectStatisticSet')
    object_statistics = object_statistics_set.findall('.//MVDObjectStatistic')
    sorted_object_statistics = sorted(object_statistics, key=lambda x: x.find('ObjectName').text)

    # Replace existing MVDObjectStatistic elements with sorted ones
    for statistic in sorted_object_statistics:
        object_statistics_set.remove(statistic)
        object_statistics_set.append(statistic)

    # Sort MVDFullStatistic elements based on InstanceName
    full_statistics_set = root.find('.//MVDFullStatisticSet')
    full_statistics = full_statistics_set.findall('.//MVDFullStatistic')
    sorted_full_statistics = sorted(full_statistics, key=lambda x: x.find('InstanceName').text)

    # Replace existing MVDFullStatistic elements with sorted ones
    for statistic in sorted_full_statistics:
        full_statistics_set.remove(statistic)
        full_statistics_set.append(statistic)

    # Sort MVDMessage elements based on MessageCode, GUID, MessageString, and optional substring
    def message_sort_key(elem):
        message_code = elem.find('MessageCode')
        message_code_text = message_code.text if message_code is not None else ''

        guid = elem.find('GUID')
        guid_text = guid.text if guid is not None else ''

        message_string = elem.find('Messages/MessageString')
        message_string_text = message_string.text if message_string is not None else ''

        instance = elem.find('Instance')
        instance_text = instance.text if instance is not None else ''

        # Extract optional substring if available
        optional_substring = ''
        if instance_text:
            start_index = instance_text.find("('") + 2
            end_index = instance_text.find("',")
            if start_index != -1 and end_index != -1:
                optional_substring = instance_text[start_index:end_index]

        if optional_substring != '':
            object_instance = elem.find('ObjectInstance')
            instance_text = object_instance.text if object_instance is not None else ''

            # Extract optional substring if available
            optional_substring = ''
            if instance_text:
                start_index = instance_text.find("('") + 2
                end_index = instance_text.find("',")
                if start_index != -1 and end_index != -1:
                    optional_substring = instance_text[start_index:end_index]

        return (message_code_text, guid_text, message_string_text, optional_substring)

    message_set = root.find('.//MVDMessageSet')
    messages = message_set.findall('.//MVDMessage')
    sorted_messages = sorted(messages, key=message_sort_key)

    # Replace existing MVDMessage elements with sorted ones
    for message in sorted_messages:
        if translation_hash is not None:
            message_code = message.findall('.//MessageCode')
            if message_code is not None and len(message_code)>0:
                try:
                    error_codes.add(message_code[0].text)
                    translation = translation_hash[message_code[0].text]
                    if translation is not None and translation != '':
                        message_code[0].text = message_code[0].text # + ": " + translation['DE'] #TODO The last part must be uncommented in order to add further documentation to sorted chkxml files
                except KeyError:
                    missing_messages.add(message_code[0].text)
        message_set.remove(message)
        message_set.append(message)

    if missing_messages:
        print("Some error codes are missing in the translation file!")
        print("Please report the following list back to the authors at https://github.com/bimhelden/basic-ifc-testing:")
        for missing in missing_messages:
            print(missing)
        print("")

    # Export statistics
    if download_csv is True:
        update_statistics('chk-statistics.json', error_codes, missing_messages)

    # Write the sorted XML to a new file
    if not output_file.endswith(".xlsx"):
        try:
            tree.write(output_file, encoding='utf-8', xml_declaration=True)
            print("ChkXML file sorted and saved as", output_file)
        except PermissionError:
            print(f"Error: Not allowed to write '{output_file}'.")
            print(f"       The file might be locked by an other application.")
            sys.exit(-1)
        except Exception as e:
            print(f"An unexpected error occurred when trying to export the sorted chkxml file: {e}")
            sys.exit(-1)

    # Export sorted messages to Excel
    excel_file = output_file.replace('.chkxml', '.xlsx')
    export_to_excel(sorted_messages, excel_file, translation_hash)
    print("Excel file saved as", excel_file)

def export_with_stylesheet(tree, output_file):
    with open(output_file, 'w', encoding='utf-8') as f:
        # Write XML declaration and root element
        f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        f.write('<?xml-stylesheet type="text/xsl" href="./MVDChecker.xsl"?>\n')

        f.write('<MVDReport>\n')

        # Write the sorted content (excluding the header)
        for child in tree.getroot():
            f.write(ET.tostring(child, encoding='unicode'))

        # Close root element
        f.write('</MVDReport>\n')

def export_to_excel(messages, output_file, translation_hash):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MVDMessages"

    # Define the headers
    headers = ["Nachrichtentyp", "Fehlerart", "Quelle", "Regel", "GUID", "Instance",
               "Instanz-Link", "Objekt", "Objekt-Link", "Description",
               "Beschreibung", "Weitere Anweisungen"]
    ws.append(headers)

    # Apply bold font to the first row (headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Iterate through the messages and add to the worksheet
    for message in messages:
        additional_content = {
            "Description": "",
            "Beschreibung": "",
            "Typ": "",
            "Instruktionen": ""
        }

        if translation_hash is not None:
            message_code = message.find('MessageCode')
            if message_code is not None:
                try:
                    translation = translation_hash[message.find('MessageCode').text]
                    if translation is not None and translation != '':
                        additional_content = {
                            "Description": translation['EN'],
                            "Beschreibung": translation['DE'],
                            "Typ": translation['TYPE'],
                            "Instruktionen": translation['INSTRUCTIONS']
                        }
                except KeyError:
                    pass # set of missing codes already handled in the sort functionality
        row = [
            message.find('MessageType').text if message.find('MessageType') is not None else '',
            additional_content["Typ"],
            message.find('MVDType').text if message.find('MVDType') is not None else '',
            message.find('MessageCode').text if message.find('MessageCode') is not None else '',
            message.find('GUID').text if message.find('GUID') is not None else '',
            message.find('Instance').text if message.find('Instance') is not None else '',
            message.find('InstanceLink').text if message.find('InstanceLink') is not None else '',
            message.find('ObjectInstance').text if message.find('ObjectInstance') is not None else '',
            message.find('ObjectInstanceLink').text if message.find('ObjectInstanceLink') is not None else '',
            additional_content["Description"],
            additional_content["Beschreibung"],
            additional_content["Instruktionen"]
        ]
        ws.append(row)

    # Add filter to all columns
    ws.auto_filter.ref = ws.dimensions

    # Adjust the width of columns B, D, and E to fit the content
    for col in ['B', 'D', 'E', 'K', 'L']:
        max_length = 0
        column = ws[col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col].width = adjusted_width

    # Hide columns except specific ones
    columns_to_hide = {"C", "F", "G", "H", "I", "J"}
    for idx, col in enumerate(headers, 1):
        if get_column_letter(idx) not in columns_to_hide:
            ws.column_dimensions[get_column_letter(idx)].hidden = False
        else:
            ws.column_dimensions[get_column_letter(idx)].hidden = True

    # Save the workbook
    try:
        wb.save(output_file)
    except PermissionError:
        print(f"Error: Not allowed to write '{output_file}'.")
        print(f"       The file might be locked by an other application.")
        sys.exit(-1)
    except Exception as e:
        print(f"An unexpected error occurred when trying to export the xlsx file: {e}")
        sys.exit(-1)

def process_options(options):
    global replace_id, download_csv, output_file
    for option in options:
        if option == "-r":
            replace_id = True
        elif option == "-u":
            download_csv = True
        else:
            output_file = option

if __name__ == "__main__":
    print("Script for creating an Excel spreadsheet with more detailed error messages from a ChkXML file created by the ifcCheckingTool by KIT (www.iai.kit.edu/ifc)")
    print("  Usage:  ChkXML2XLSX.exe input_file <output_file> <option>")
    print("  Option: -u download latest definitions and upload basic statistics")
    print("          -r will replace spf ids by #uid")
    print("developed by AEC3 Germany (www.aec3.de), use on own risk.")
    print("Further information can be found at: https://github.com/bimhelden/basic-ifc-testing")

    if len(sys.argv) == 1:
        # Prompt user for input and output file names, and further options
        input_file = input("\nEnter the input file name: ")
        output_file = input("Enter the output file name: ")
        replace_id = input("Replace SPF-IDs (y/n): ").lower() == 'y'
        download_csv = input("Download latest translation file and upload basic statistics (y/n): ").lower() == 'y'
    elif len(sys.argv) in [2, 3, 4, 5]:
        input_file = sys.argv[1]
        options = sys.argv[2:]
        process_options(options)
    else:
        print(f"Unexpected number of arguments: '{sys.argv}'")
        sys.exit(-1)

    print("")

    # read file with translation of WRs
    csv_filename = "ChkXML-IFC4Add2TC1-DE.csv"
    if download_csv is True:
        try:
            fetch_file_from_github(csv_filename)
        except Exception as e:
            print(f"Download of the translation file failed with the following error code: {e}")

    wr_translation = read_csv_to_hash(csv_filename)

    if output_file == "":
        output_file = input_file.replace('.chkxml', '.xlsx')
    try:
        sort_xml(input_file, output_file, replace_id, wr_translation, download_csv)
    except Exception as e:
        print(f"An unexpected error occurred when trying to process the chkxml file: {e}")
        sys.exit(-1)

